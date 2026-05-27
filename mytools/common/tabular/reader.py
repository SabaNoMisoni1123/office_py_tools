from __future__ import annotations

import csv
from pathlib import Path
from typing import Any


def read_table(
    input_path: Path,
    *,
    sheet_name: str | None,
    encoding: str,
    header_row: int,
) -> list[dict[str, Any]]:
    suffix = input_path.suffix.lower()
    if suffix == ".csv":
        if sheet_name is not None:
            raise ValueError("CSV 入力では --sheet を指定できません。")
        return read_csv_table(input_path, encoding=encoding, header_row=header_row)
    if suffix == ".xlsx":
        return read_xlsx_table(input_path, sheet_name=sheet_name, header_row=header_row)
    raise ValueError("入力ファイルは .csv または .xlsx を指定してください。")


def read_csv_table(
    input_path: Path, *, encoding: str, header_row: int
) -> list[dict[str, Any]]:
    with input_path.open("r", encoding=encoding, newline="") as f:
        rows = list(csv.reader(f))
    return rows_to_dicts(rows, header_row=header_row)


def read_xlsx_table(
    input_path: Path, *, sheet_name: str | None, header_row: int
) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError as exc:
        raise ImportError(
            "xlsx 入力を読むには openpyxl が必要です。依存関係をインストールしてください。"
        ) from exc

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        if sheet_name is None:
            sheet = workbook[workbook.sheetnames[0]]
        else:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"指定シートが見つかりません: {sheet_name}")
            sheet = workbook[sheet_name]
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        return rows_to_dicts(rows, header_row=header_row)
    finally:
        workbook.close()


def rows_to_dicts(rows: list[list[Any]], *, header_row: int) -> list[dict[str, Any]]:
    if header_row < 1:
        raise ValueError("header_row は 1 以上を指定してください。")
    header_index = header_row - 1
    if len(rows) <= header_index:
        raise ValueError("ヘッダー行が入力データの範囲外です。")

    headers = [normalize_header(value) for value in rows[header_index]]
    if not any(headers):
        raise ValueError("ヘッダー行が空です。")
    if len(headers) != len(set(headers)):
        raise ValueError("ヘッダー行に重複した列名があります。")

    records: list[dict[str, Any]] = []
    for row in rows[header_index + 1 :]:
        if not any(value not in (None, "") for value in row):
            continue
        record: dict[str, Any] = {}
        for index, header in enumerate(headers):
            record[header] = row[index] if index < len(row) else None
        records.append(record)
    return records


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()

